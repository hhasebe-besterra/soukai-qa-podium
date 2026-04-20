"""一般想定問答xlsxから Q&A を抽出して general_qa.json を生成する (1回だけ実行)"""
import sys, re, json, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

XLSX_PATH = r'C:\Users\h.hasebe\Downloads\20260413 最新 ■想定問答1-22（近藤修正）.xlsx'
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'general_qa.json')

SHEET_PAT = re.compile(r'^[！]?(\d+)-(\d+)(?:\s*\(補足\)|\s*補足)?$')

# PDF 表紙に基づく正本カテゴリ名（xlsxセル抽出は項目セル位置ズレの可能性あり）
CANONICAL_CATS = {
    1: '外部環境', 2: '株主総会', 3: '役員', 4: '監査等委員会', 5: '顧問',
    6: 'プライム', 7: '資本政策', 8: '配当', 9: '経営戦略',
    10: '関係会社・M&A', 11: '財務諸表', 12: '設備投資', 13: '営業活動',
    14: '地方事務所', 15: '特許', 16: '技術', 17: '工事', 18: 'DX',
    19: '採用', 20: '人事', 21: 'リスク管理', 22: '内部統制',
}

def extract_texts(ws):
    """シート内の全テキスト（結合セル含む）を行単位で取得"""
    rows_text = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [str(c).strip() for c in row if c and str(c).strip()]
        rows_text.append(cells)
    return rows_text

def parse_sheet(ws):
    """質問番号, 項目, 回答者, 質問事項, 回答 を抽出"""
    rows = extract_texts(ws)
    data = {'no': '', 'item': '', 'responder': '', 'q': '', 'a': ''}
    # header line
    for i, cells in enumerate(rows[:5]):
        joined = ' | '.join(cells)
        if '質問番号' in joined:
            # look for hyphen-separated no like "1-1"
            for c in cells:
                if re.match(r'^\d+-\d+', c):
                    data['no'] = c
            # 項目
            for j, c in enumerate(cells):
                if c == '項目' and j+1 < len(cells):
                    data['item'] = cells[j+1]
            # 回答者
            for j, c in enumerate(cells):
                if c == '回答者' and j+1 < len(cells):
                    data['responder'] = cells[j+1]
    # question + answer: scan for label 質問事項 / （回答）
    mode = None
    q_buf = []
    a_buf = []
    for cells in rows:
        joined = ''.join(cells)
        if not joined:
            if mode == 'a' and a_buf:
                continue  # keep accumulating answer over blank rows
            continue
        if '質問事項' in joined and len(joined) < 20:
            mode = 'q'; continue
        if '回' in joined and '答' in joined and len(joined) < 15:
            mode = 'a'; continue
        if mode == 'q':
            q_buf.append(' '.join(cells))
        elif mode == 'a':
            a_buf.append(' '.join(cells))
    data['q'] = '\n'.join(q_buf).strip()
    data['a'] = '\n'.join(a_buf).strip()
    return data

def main():
    print(f'loading {XLSX_PATH}')
    wb = load_workbook(XLSX_PATH, data_only=True)
    items = []
    cat_labels = {}
    for sn in wb.sheetnames:
        m = SHEET_PAT.match(sn)
        if not m:
            continue
        major = int(m.group(1))
        minor = int(m.group(2))
        is_supp = '補足' in sn
        ws = wb[sn]
        try:
            d = parse_sheet(ws)
        except Exception as e:
            print(f'  ! parse error {sn}: {e}')
            continue
        if not d['q']:
            print(f'  - skip {sn} (no question)')
            continue
        display_id = d['no'] or f'{major}-{minor}'
        if is_supp:
            display_id += '補'
        canon_label = CANONICAL_CATS.get(major, d['item'] or str(major))
        item = {
            'id': display_id,
            'major': major,
            'minor': minor,
            'cat': str(major),
            'catLabel': f"{major}. {canon_label}",
            'q': d['q'],
            'a': d['a'] or '(回答未記載)',
            'responder': d['responder'],
            'tag': 'answered' if d['a'] else 'declined',
            'src': f'想定問答 rev2026-04-13 / {sn}',
        }
        items.append(item)

    # categories (ordered by major)
    cats = []
    for k, lbl in CANONICAL_CATS.items():
        n = sum(1 for it in items if it['major']==k)
        if n > 0:
            cats.append([str(k), f"{k}. {lbl}", n])

    out = {'items': items, 'cats': cats}
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f'\nsaved {OUT_PATH}')
    print(f'items: {len(items)}')
    print(f'categories: {len(cats)}')
    for c in cats: print(f'  {c[0]:>3}: {c[1]} ({c[2]})')

if __name__ == '__main__':
    main()
